import pandas as pd
import numpy as np
import openpyxl
import xlrd
import os
import logging
from ClassProject.CloseWs import CloseWs


class FailDataExcel(Exception):
    def __init__(self, msg):
        self.msg=msg
    
    def __str__(self):
        return self.msg

class DataExcel:
    """
    La clase `DataExcel` se encarga de gestionar la carga de un archivo de Excel en formato .xls o .xlsx.
    Los archivos de excel trabajados serán transformados todos a formato texto para evitar errores de tipo de dato.
    Los dataframes generados por esta clase serán de tipo `pandas.DataFrame` pero con dtype `object` en todas las columnas.

    Atributos:
        path (str): Ruta del archivo Excel a cargar.
        header (str): Nombre de la cabecera a utilizar para identificar el inicio de los datos en el archivo.
        _frame (DataFrame): DataFrame resultante de cargar los datos del archivo.

    Métodos:
        path_verify() -> Verifica la ruta del archivo y la transforma a formato .xlsx en caso de ser necesario.
        load_excel() -> Carga los datos del archivo a un DataFrame.
        transform_xlsx_values_to_text(file_path:str) -> Transforma los valores de todas las celdas de un archivo .xlsx a formato de texto.
        transform_xls_to_xlsx(file_path:str) -> Transforma un archivo .xls a .xlsx y transforma los valores de todas las celdas a formato de texto.
    """
    def __init__(self, path:str, header:str = None, dtype_str:bool = False):
        """
         Inicializa la clase `DataExcel`.

        Args:
            path (str): Ruta del archivo Excel a cargar.
            header (str, opcional): Nombre de la cabecera a utilizar para identificar el inicio de los datos en el archivo. Por defecto, None.
            dtype_str(bool, opcional): Especifica si se quiere que todo el dataframe se lea como celdas con formato string. Por defecto, False,
                lo que indica que deja que el establecimientos de dtypes dependerá de pd.read_excel().
        """
        self._path=path
        self.path=self.path_verify()
        self.header=header
        self.dtype_str= dtype_str
        self._frame=None   
    
    @property
    def frame(self) -> pd.DataFrame:
        """
        Propiedad que permite acceder al DataFrame generado a partir del archivo de Excel.

        Returns
        -------
        pandas.DataFrame
            El DataFrame generado a partir del archivo de Excel.
        """
        self._frame=self.load_excel()
        return self._frame
    
    
    def path_verify(self):
        """
        Verifica la ruta del archivo y la transforma a formato .xlsx en caso de ser necesario.

        Returns:
            str: Ruta del archivo verificada y transformada.

        Raises:
            FailDataExcel: Si se presenta algún error durante la verificación de la ruta.
        """
        try:
            path=self._path
            
            if path.endswith(".xls"):
                try:
                    path_xlsx=path.replace(".xls",".xlsx")
                    DataExcel.transform_xls_to_xlsx(path)
                    return path_xlsx
                
                except xlrd.XLRDError:
                    path_xlsx=path.replace(".xls",".xlsx")
                    if os.path.exists(path_xlsx):
                        os.remove(path_xlsx)
                    os.rename(path,path_xlsx)
                    DataExcel.transform_xlsx_values_to_text(path_xlsx)
                    return path_xlsx
            
            elif path.endswith(".xlsx"):
                DataExcel.transform_xlsx_values_to_text(path)
                return path
            
            else:
                logging.error("El archivo no es de tipo xlsx o xls")
                raise TypeError("El archivo no tiene formato válido .xlsx o .xls")
        
        
        except Exception as e:
            logging.error("DataExcel : Problema al ejecutar path_verify()")
            logging.error(f"{e.__class__}: {e}")
            raise FailDataExcel(f"Error en la verificación del path: {e.__class__}: {e}")
                    
    
    def load_excel(self):
        """
        Método que carga los datos del archivo xlsx y retorna un DataFrame con todas las columnas de tipo `object`
        en caso el argumento dtype_str tenga valor True al inicializar el objeto DataExcel.
        Por defecto Pandas interpretará el tipo de dato en cada columna

        Returns:
            pandas.DataFrame: El DataFrame generado a partir del archivo de Excel.

        Raises:
            FailDataExcel: Si se presenta algún error durante la carga de los datos.
        """
        read_as_string = str if self.dtype_str is True else None
        try:
            
            frame=pd.read_excel(self.path, header=None, dtype=read_as_string)
            if self.header is not None:
                head=frame[frame.eq(self.header).any(axis=1, bool_only=None)].index.values
                if head.size > 0:
                    return pd.read_excel(self.path, header=head[0], dtype=read_as_string)
            
            return pd.read_excel(self.path, dtype=read_as_string)
        
        except Exception as e:
            logging.error("DataExcel : Problema al ejecutar load_excel()")
            logging.error(f"{e.__class__}: {e}")
            raise FailDataExcel(f"Error al ejecutar load_excel(): {e.__class__}: {e}")
        
    
    @staticmethod
    def transform_xlsx_values_to_text(file_path):
        """
        Transforma los valores de todas las celdas de un archivo .xlsx a formato de texto.

        Args:
            file_path (str): Ruta del archivo .xlsx a transformar.

        Returns:
            bool: True si la transformación se realiza con éxito, False en caso contrario.

        Raises:
            FailDataExcel: Si se presenta algún error durante la transformación.
        """
        wb = openpyxl.load_workbook(file_path)
        sheet=wb.active
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    cell.value=""
                else:
                    cell.value = str(cell.value)
                cell.number_format = "@"
        wb.save(file_path)
        wb.close()
        return True
    
    @staticmethod
    def transform_xls_to_xlsx(file_path):
        """
        Transforma un archivo .xls a .xlsx y transforma los valores de todas las celdas a formato de texto.

        Args:
            file_path (str): Ruta del archivo .xls a transformar.

        Returns:
            bool: True si la transformación se realiza con éxito, False en caso contrario.

        Raises:
            FailDataExcel: Si se presenta algún error durante la transformación.
        """
        wb=xlrd.open_workbook(file_path)
        wb_xlsx = openpyxl.Workbook()
        path_xlsx=file_path.replace(".xls",".xlsx")
        
        sh = wb.sheet_by_index(0)
        sh_destino = wb_xlsx.active
        
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                valor=sh.cell_value(r,c)
                if valor is None:
                    sh_destino.cell(row=r+1, column=c+1).value = ""
                else:
                    sh_destino.cell(row=r+1, column=c+1).value = str(valor)
                sh_destino.cell(row=r+1, column=c+1).number_format = "@"
        
        if os.path.exists(path_xlsx):
            os.remove(path_xlsx)                
        
        wb_xlsx.save(path_xlsx)
        wb_xlsx.close()
        wb.close()
        os.remove(file_path)
        return True